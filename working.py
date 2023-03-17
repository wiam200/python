from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
import math

wb=load_workbook('modele.xlsx')
ws=wb.active

# The firat one
cout_sans_raccordement = 4499
print('------------------------------------------------------------------------------------------------------------------------')

print("hello fucking world")
print("hello fucking world")

print("hello fucking world")

print()
print()
print()

distance_installe=input('Entrer la puissance à installer en (MW): ')
i_sans_raccoredement = (4499/1000000)*(int(distance_installe)*1000)

# The second one
cout_unitaire_PGA=50035
distance_PGA=input('Entrer la distance de raccordement des lignes éléctriques haute-tension HT en (KM):')
cout_raccordement_PGA=(cout_unitaire_PGA/1000000)*distance_PGA

#The 3rd one
cout_unitaire_GR=2000000
distance_GR=input('Entrer la distance e raccordement des ressources en gaz < pipeline > en (KM):')
cout_raccordement_GR=(cout_unitaire_GR/1000000)*int(distance_GR)

#The 4rd one
cout_unitaire_WA=292215
distance_WA=input('Entrer la distance de raccordement à l`eau < réseau de distribution ou barrage > en (KM):')
cout_raccordement_WA=(cout_unitaire_WA/1000000)*int(distance_WA)

#The 5rd one
cout_unitaire_RA=3960396
distance_RA=input('Entrer la distance de raccordement à la route en (KM):')
cout_raccordement_RA=(cout_unitaire_RA/1000000)*int(distance_RA)

# Le cout total

cout_total_des_raccordements=cout_raccordement_PGA+cout_raccordement_GR+cout_raccordement_RA+cout_raccordement_WA

# Investissement

Investissement=cout_total_des_raccordements+i_sans_raccoredement

# O&M

O_and_M=Investissement*0.025

# DNI

DNI=input('Donner le rayonnement solaire direct DNI (kWh/m2/an)  :')
SM=input('Donner le Solar Multiple (=2 : avec stockage ; 1 seul solar field ; 7.5 h de stockage) :')

# FLH

FLH=(2.5717*int(DNI)+694)*(-0.0371*int(SM)*2+0.4171*int(SM)-0.0744)

# EY
Ey= (int(distance_installe)/1000)*FLH

# i , n
i=0.08
n=25

# LCEO sans raccordement

LCOE_sans_raccordement=(i_sans_raccoredement*(math.pow((1+i),n)*i/(math.pow((1+i),n)-1))+O_and_M)/Ey*100

# % du cout de raccordement / I
pourcentage_du_cout_sur_I=LCOE_sans_raccordement/Investissement*100

# LCOE

LCOE=(Investissement*(math.pow((1+i),n)*i/(math.pow(1+i,n)-1))+O_and_M)/Ey

#LCOE (cUSD/kwh)
LCOE_cUSD=LCOE*100

# LCOE (DA/KWH)
LCOE_DA=LCOE_cUSD*140.45/100

#AFFICHAge
print()
print('-------- LES RESULTAS --------')
print()

print('Le Coût d`investissement de la centrale sans raccoredement (Million USD) = ',round(i_sans_raccoredement,2))

print()

print('Le cout de raccordement des lignes éléctriques haute-tension HT (Million USD) =',round(cout_raccordement_PGA,2))
print('Le cout de raccordement des ressources en gaz < pipeline > (Million USD) =',round(cout_raccordement_GR,2))
print('Le cout de raccordement à l`eau < réseau de distribution ou barrage > (Million USD) =',round(cout_raccordement_WA,2))
print('Le cout de raccordement à la route (Million USD) =',round(cout_raccordement_RA,2))

print()

print('Le cout des raccordements (Million USD) = ',round(cout_total_des_raccordements,2))

print()

print('----------')

print()
print('L`investissement (Million USD) =',round(Investissement,2))
print('O&M Les coûts annuels d`exploitation et de maintenance (Million USD) = ',round(O_and_M,2))

print()

print('----------')

print()


print('FLH Les heures de fonctionnement à pleine charge solaire (heures) = ',round(FLH,2))
print('L`Energie électrique produite par la centrale CSP en un an (GWH/AN) = ', round(Ey,2))
print()

print('----------')
print()

print('<LCOE> le cout de production de l`électricité sans raccordement (USD/kWh) = ',round(LCOE_sans_raccordement,2))
print(' % du cout de raccordement sur I = ',round(pourcentage_du_cout_sur_I,2))

print()

print('<LCOE> le cout de production de l`électricité (USD/kWh)  = ',round(LCOE,3))

print()

print('<LCOE> le cout de production de l`électricité (cUSD/kwh) = ',round(LCOE_cUSD,3))

print()

print('<LCOE> le cout de production de l`électricité (DA/kwh) = ',round(LCOE_DA,2))

print()
print()
print()